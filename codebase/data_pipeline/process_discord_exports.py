from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta
from difflib import SequenceMatcher
import hashlib
import json
import os
from pathlib import Path
import re
import statistics
import unicodedata
from urllib.parse import unquote, urlsplit, urlunsplit
from urllib.request import Request, urlopen


SPACE_RE = re.compile(r"\s+")
URL_RE = re.compile(r"https?://[^\s,]+", re.IGNORECASE)
EMAIL_RE = re.compile(r"(?<![\w.-])[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}(?![\w.-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?84|0)(?:[ .-]?\d){8,10}(?!\d)")
DOT_NOISE_RE = re.compile(r"^[\s.…·•,;:_-]+$")
QUESTION_RE = re.compile(
    r"(?:\?|\b(?:ai|gì|sao|nào|đâu|bao giờ|khi nào|có (?:ai|cách)|làm thế nào|"
    r"how|what|why|where|when|help)\b)",
    re.IGNORECASE,
)
PROBLEM_RE = re.compile(
    r"\b(?:lỗi|error|bug|issue|problem|vướng|kẹt|blocker|không (?:được|thấy|nhận|vào|hiện)|"
    r"chưa (?:được|có|nhận|thấy)|invalid|failed|failure|not a member|can't|cannot|mất|"
    r"bị (?:lỗi|kẹt|chặn)|cứu|hỗ trợ|giúp)\b",
    re.IGNORECASE,
)
REQUEST_RE = re.compile(
    r"\b(?:nhờ|xin|cho (?:mình|em)|giúp|hỗ trợ|hướng dẫn|làm sao|cách nào|please)\b",
    re.IGNORECASE,
)
ADVICE_RE = re.compile(
    r"^(?:bạn|em|mọi người)?\s*(?:hãy|thử|dùng|kiểm tra|vào|gõ|tạo|đợi|chờ)\b",
    re.IGNORECASE,
)

GREETINGS = {
    "hi", "hello", "hey", "xin chào", "chào", "chào mọi người", "good day",
    "alo", "hế lô", "hé lô", "ciao", "morning", "buổi sáng tốt lành",
}
ACKNOWLEDGEMENTS = {
    "ok", "okay", "oke", "oki", "dạ", "vâng", "rõ", "được rồi", "xong rồi",
    "cảm ơn", "cám ơn", "thanks", "thank you", "done", "đã hiểu", "em hiểu rồi",
}
BOT_MARKERS = ("trợ lý kute", "#3191", " bot", "bot-")

CHANNEL_BY_FILE = {
    "crawl đoạn chat chung.xlsx": "common",
    "crawl đoạn chat chung.xlsx": "common",
    "crawl_command.xlsx": "bot-commands",
    "crawl_qa_discord.xlsx": "qa",
}

ENTITY_PATTERNS = {
    "GitHub": ("github", "organization", "org invite", "lời mời org"),
    "VLearn": ("vlearn", "v-learn"),
    "Discord": ("discord", "channel", "role", "server"),
    "Onboarding": ("onboarding", "invite", "lời mời", "tham gia"),
    "Slide": ("slide", "bài giảng", "lecture"),
    "Lab": ("lab", "thực hành", "d304"),
    "Assignment": ("bài tập", "assignment", "deadline", "nộp bài"),
    "Account": ("đăng nhập", "login", "credential", "tài khoản", "password"),
    "Attendance": ("xin nghỉ", "điểm danh", "vắng", "attendance"),
}


@dataclass
class EpisodeDraft:
    anchor_index: int
    message_indexes: list[int]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFC", str(value))
    return SPACE_RE.sub(" ", text).strip()


def searchable_text(value: str) -> str:
    return normalize_text(value).casefold()


def redact_for_model(value: str) -> str:
    redacted = EMAIL_RE.sub("[EMAIL_REDACTED]", value)
    return PHONE_RE.sub("[PHONE_REDACTED]", redacted)


def stable_key(*parts: object, length: int = 16) -> str:
    payload = "|".join(str(part) for part in parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length]


def normalized_author_id(value: object) -> str:
    if value is None or value == "":
        return "unknown"
    if isinstance(value, float):
        return format(value, ".0f")
    return normalize_text(value)


def strip_url_token(url: str) -> str:
    parts = urlsplit(url)
    return urlunsplit((parts.scheme.casefold(), parts.netloc.casefold(), parts.path, "", ""))


def parse_attachments(value: object) -> list[dict]:
    raw = normalize_text(value)
    attachments: list[dict] = []
    for match in URL_RE.findall(raw):
        original = match.rstrip(")]}>")
        canonical = strip_url_token(original)
        filename = unquote(Path(urlsplit(canonical).path).name)
        extension = Path(filename).suffix.casefold().lstrip(".")
        attachments.append(
            {
                "url": original,
                "filename": filename,
                "extension": extension,
                "caption_ocr": None,
                "canonical_url": canonical,
            }
        )
    return attachments


def parse_reactions(value: object) -> tuple[list[dict], int]:
    raw = normalize_text(value)
    if not raw:
        return [], 0
    reactions: list[dict] = []
    for part in re.split(r",(?=\s*\S)", raw):
        match = re.match(r"^\s*(.*?)\s*\((\d+)\)\s*$", part)
        if not match:
            continue
        reactions.append({"emoji": match.group(1), "count": int(match.group(2))})
    return reactions, sum(item["count"] for item in reactions)


def _short_phrase_match(value: str, phrases: set[str]) -> bool:
    compact = value.strip(" .,!?:;~-_()[]{}").casefold()
    return compact in phrases or (
        len(compact) <= 28 and any(compact.startswith(f"{phrase} ") for phrase in phrases)
    )


def classify_flags(content: str, author_name: str, attachments: list[dict]) -> dict:
    lowered = searchable_text(content)
    is_dot_noise = bool(content and DOT_NOISE_RE.fullmatch(content))
    is_greeting = _short_phrase_match(lowered, GREETINGS)
    is_acknowledgement = _short_phrase_match(lowered, ACKNOWLEDGEMENTS) or any(
        phrase in lowered for phrase in ("cảm ơn", "cám ơn", "thank", "đã được", "ổn rồi")
    )
    is_bot = any(marker in searchable_text(author_name) for marker in BOT_MARKERS)
    has_attachment = bool(attachments)
    is_attachment_only = has_attachment and (not content or is_dot_noise)
    return {
        "is_dot_noise": is_dot_noise,
        "is_greeting": is_greeting,
        "is_acknowledgement": is_acknowledgement,
        "is_bot": is_bot,
        "is_question": bool(QUESTION_RE.search(lowered)),
        "is_problem": bool(PROBLEM_RE.search(lowered)),
        "has_attachment": has_attachment,
        "is_attachment_only": is_attachment_only,
    }


def discover_sheet(workbook):
    return max(workbook.worksheets, key=lambda sheet: sheet.max_row)


def read_exports(input_dir: Path) -> list[dict]:
    from openpyxl import load_workbook

    messages: list[dict] = []
    for path in sorted(input_dir.glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = discover_sheet(workbook)
        rows = sheet.iter_rows(values_only=True)
        headers = [normalize_text(value) for value in next(rows)]
        header_map = {name.casefold(): index for index, name in enumerate(headers)}
        required = {"authorid", "author", "date", "content", "attachments", "reactions"}
        if not required.issubset(header_map):
            continue
        channel_key = CHANNEL_BY_FILE.get(path.name.casefold(), path.stem.casefold())
        for source_row, row in enumerate(rows, start=2):
            def cell(name: str):
                index = header_map[name]
                return row[index] if index < len(row) else None

            content_original = "" if cell("content") is None else str(cell("content"))
            content_clean = normalize_text(content_original)
            attachments_original = "" if cell("attachments") is None else str(cell("attachments"))
            reactions_original = "" if cell("reactions") is None else str(cell("reactions"))
            attachments = parse_attachments(attachments_original)
            reactions, reaction_count = parse_reactions(reactions_original)
            author_id = normalized_author_id(cell("authorid"))
            author_name = normalize_text(cell("author")) or "Unknown"
            date_value = cell("date")
            if isinstance(date_value, datetime):
                created_at = date_value.isoformat()
            else:
                created_at = normalize_text(date_value)
            content_search = searchable_text(content_clean)
            flags = classify_flags(content_clean, author_name, attachments)
            message_key = f"{channel_key}-{source_row}"
            messages.append(
                {
                    "message_key": message_key,
                    "source_file": path.name,
                    "source_sheet": sheet.title,
                    "source_row": source_row,
                    "channel_key": channel_key,
                    "author_id": author_id,
                    "reporter_key": f"reporter-{stable_key(author_id)}",
                    "author_name": author_name,
                    "created_at": created_at,
                    "content_original": content_original,
                    "content_clean": content_clean,
                    "content_search": content_search,
                    "content_model": redact_for_model(content_clean),
                    "attachments_original": attachments_original,
                    "attachments": attachments,
                    "reactions_original": reactions_original,
                    "reactions": reactions,
                    "reaction_count": reaction_count,
                    "exact_normalized_hash": (
                        stable_key(content_search, length=20) if content_search else None
                    ),
                    "same_author_duplicate_of": None,
                    "episode_message_group_id": None,
                    **flags,
                }
            )
    messages.sort(key=lambda item: (item["channel_key"], item["created_at"], item["source_row"]))
    mark_near_duplicates(messages)
    return messages


def parse_datetime(value: str) -> datetime:
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return datetime.min


def text_similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    return SequenceMatcher(None, left, right).ratio()


def mark_near_duplicates(messages: list[dict]) -> None:
    latest_by_author: dict[tuple[str, str], list[int]] = {}
    for index, message in enumerate(messages):
        key = (message["channel_key"], message["reporter_key"])
        created = parse_datetime(message["created_at"])
        candidates = latest_by_author.setdefault(key, [])
        candidates[:] = [
            position
            for position in candidates
            if created - parse_datetime(messages[position]["created_at"]) <= timedelta(minutes=20)
        ]
        duplicate_index = next(
            (
                position
                for position in reversed(candidates)
                if text_similarity(message["content_search"], messages[position]["content_search"]) >= 0.9
            ),
            None,
        )
        if duplicate_index is not None:
            duplicate = messages[duplicate_index]
            message["same_author_duplicate_of"] = duplicate["message_key"]
            message["episode_message_group_id"] = (
                duplicate["episode_message_group_id"]
                or f"msg-group-{stable_key(key, duplicate['message_key'])}"
            )
            duplicate["episode_message_group_id"] = message["episode_message_group_id"]
        candidates.append(index)


def extract_entities(text: str) -> list[str]:
    lowered = searchable_text(text)
    return [
        entity
        for entity, patterns in ENTITY_PATTERNS.items()
        if any(pattern in lowered for pattern in patterns)
    ]


def product_area(entities: list[str], text: str) -> str:
    values = set(entities)
    lowered = searchable_text(text)
    if "GitHub" in values:
        return "developer_access"
    if "VLearn" in values and "Account" in values:
        return "vlearn_access"
    if "VLearn" in values:
        return "learning_platform"
    if "Attendance" in values:
        return "attendance"
    if "Assignment" in values:
        return "coursework"
    if "Discord" in values or "Onboarding" in values:
        return "community_onboarding"
    if "Slide" in values or "Lab" in values:
        return "learning_content"
    if any(value in lowered for value in ("mail", "account", "tài khoản", "đăng nhập")):
        return "account_access"
    return "other"


def useful_for_context(message: dict) -> bool:
    return not message["is_dot_noise"] and not message["is_greeting"] and bool(
        message["content_model"] or message["attachments"]
    )


def is_anchor(message: dict) -> bool:
    if message["is_bot"] or message["is_dot_noise"] or message["is_greeting"]:
        return False
    if (
        ADVICE_RE.search(message["content_search"])
        and not message["is_question"]
        and not REQUEST_RE.search(message["content_search"])
    ):
        return False
    described_attachment = message["has_attachment"] and not message["is_attachment_only"]
    return bool(
        message["is_question"]
        or message["is_problem"]
        or described_attachment
        or REQUEST_RE.search(message["content_search"])
    )


def relation_score(anchor: dict, candidate: dict) -> float:
    if anchor["reporter_key"] == candidate["reporter_key"]:
        return 1.0
    anchor_entities = set(extract_entities(anchor["content_model"]))
    candidate_entities = set(extract_entities(candidate["content_model"]))
    shared = anchor_entities & candidate_entities
    score = 0.42 * text_similarity(anchor["content_search"], candidate["content_search"])
    if shared:
        score += 0.52
    if candidate["is_acknowledgement"]:
        score += 0.18
    if anchor["author_name"].split("-")[-1].casefold() in candidate["content_search"]:
        score += 0.35
    return min(score, 1.0)


def build_episodes(messages: list[dict]) -> list[dict]:
    by_channel: dict[str, list[int]] = {}
    for index, message in enumerate(messages):
        by_channel.setdefault(message["channel_key"], []).append(index)

    episodes: list[dict] = []
    consumed_anchor_groups: set[str] = set()
    for channel_indexes in by_channel.values():
        for local_position, anchor_index in enumerate(channel_indexes):
            anchor = messages[anchor_index]
            if not is_anchor(anchor):
                continue
            duplicate_group = anchor["episode_message_group_id"]
            if duplicate_group and duplicate_group in consumed_anchor_groups:
                continue
            if duplicate_group:
                consumed_anchor_groups.add(duplicate_group)

            anchor_time = parse_datetime(anchor["created_at"])
            before: list[int] = []
            after: list[int] = []
            reporter_followups: list[int] = []

            for candidate_index in reversed(channel_indexes[:local_position]):
                candidate = messages[candidate_index]
                if anchor_time - parse_datetime(candidate["created_at"]) > timedelta(minutes=5):
                    break
                if useful_for_context(candidate) and relation_score(anchor, candidate) >= 0.24:
                    before.append(candidate_index)
                if len(before) >= 3:
                    break

            for candidate_index in channel_indexes[local_position + 1 :]:
                candidate = messages[candidate_index]
                delta = parse_datetime(candidate["created_at"]) - anchor_time
                if delta > timedelta(minutes=20):
                    break
                if not useful_for_context(candidate):
                    continue
                if candidate["reporter_key"] == anchor["reporter_key"]:
                    reporter_followups.append(candidate_index)
                if delta <= timedelta(minutes=10) and len(after) < 12:
                    score = relation_score(anchor, candidate)
                    if score >= 0.30:
                        after.append(candidate_index)

            context_indexes = sorted(set(before + after + reporter_followups))
            episode_messages = [anchor] + [messages[index] for index in context_indexes]
            episode_messages.sort(key=lambda item: parse_datetime(item["created_at"]))
            answers = [
                message
                for message in episode_messages
                if parse_datetime(message["created_at"]) > anchor_time
                and message["reporter_key"] != anchor["reporter_key"]
                and not message["is_acknowledgement"]
                and useful_for_context(message)
            ]
            reporter_acks = [
                message
                for message in episode_messages
                if parse_datetime(message["created_at"]) > anchor_time
                and message["reporter_key"] == anchor["reporter_key"]
                and message["is_acknowledgement"]
            ]
            first_answer_seconds = None
            if answers:
                first_answer_seconds = max(
                    0,
                    int((parse_datetime(answers[0]["created_at"]) - anchor_time).total_seconds()),
                )
            answer_summary = " ".join(item["content_model"] for item in answers[:3]).strip()
            workaround = any(
                marker in searchable_text(answer_summary)
                for marker in ("tạm", "thử", "workaround", "ticket", "hàng đợi", "limit")
            )
            if reporter_acks and answers:
                resolution_status = "resolved"
            elif answers and workaround:
                resolution_status = "workaround"
            elif answers:
                resolution_status = "unclear"
            else:
                resolution_status = "unresolved"

            canonical = anchor["content_model"][:500]
            entities = extract_entities(" ".join(item["content_model"] for item in episode_messages))
            area = product_area(entities, canonical)
            attachment_names = [
                attachment["filename"]
                for message in episode_messages
                for attachment in message["attachments"]
                if attachment["filename"]
            ]
            source_rows = [item["message_key"] for item in episode_messages]
            confidence = 0.5
            confidence += 0.15 if anchor["is_problem"] else 0
            confidence += 0.12 if anchor["is_question"] else 0
            confidence += 0.10 if entities else 0
            confidence += 0.08 if answers else 0
            episodes.append(
                {
                    "episode_id": f"episode-{stable_key(anchor['message_key'])}",
                    "reporter_key": anchor["reporter_key"],
                    "channel_key": anchor["channel_key"],
                    "start_time": episode_messages[0]["created_at"],
                    "end_time": episode_messages[-1]["created_at"],
                    "anchor_message": {
                        "message_key": anchor["message_key"],
                        "content": anchor["content_model"],
                    },
                    "context_messages": [
                        {
                            "message_key": item["message_key"],
                            "author_key": item["reporter_key"],
                            "created_at": item["created_at"],
                            "content": item["content_model"],
                            "is_acknowledgement": item["is_acknowledgement"],
                        }
                        for item in episode_messages
                        if item["message_key"] != anchor["message_key"]
                    ],
                    "canonical_problem": canonical,
                    "product_area": area,
                    "entities": entities,
                    "answer_summary": answer_summary or None,
                    "resolution_status": resolution_status,
                    "time_to_first_answer": first_answer_seconds,
                    "attachment_summary": attachment_names,
                    "reaction_count": sum(item["reaction_count"] for item in episode_messages),
                    "source_rows": source_rows,
                    "confidence": round(min(confidence, 0.95), 2),
                    "painpoint_cluster_id": None,
                }
            )
    return episodes


def semantic_text(episode: dict) -> str:
    entity_text = " ".join(episode["entities"])
    value = f"{episode['canonical_problem']} {episode['product_area']} {entity_text}"
    lowered = searchable_text(value)
    replacements = {
        "not a member of any organizations": "github organization invite membership",
        "không hiện lời mời": "không nhận invite",
        "không nhận được lời mời": "không nhận invite",
        "invalid credential": "đăng nhập lỗi credential",
        "login": "đăng nhập",
    }
    for source, target in replacements.items():
        lowered = lowered.replace(source, target)
    return lowered


def cluster_episodes(episodes: list[dict], threshold: float = 0.38) -> list[list[int]]:
    if not episodes:
        return []
    from scipy.sparse import hstack
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity

    texts = [semantic_text(episode) for episode in episodes]
    word = TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True).fit_transform(texts)
    char = TfidfVectorizer(
        analyzer="char_wb", ngram_range=(3, 5), min_df=1, sublinear_tf=True
    ).fit_transform(texts)
    vectors = hstack([word * 0.66, char * 0.34]).tocsr()
    similarity = cosine_similarity(vectors)

    clusters: list[list[int]] = []
    for index, episode in enumerate(episodes):
        best_cluster = None
        best_score = -1.0
        for cluster_index, members in enumerate(clusters):
            representative = members[0]
            same_area = episodes[representative]["product_area"] == episode["product_area"]
            shared_entities = set(episodes[representative]["entities"]) & set(episode["entities"])
            score = float(similarity[index, representative])
            required = threshold if same_area else 0.56
            if shared_entities:
                required -= 0.08
            if score >= required and score > best_score:
                best_cluster = cluster_index
                best_score = score
        if best_cluster is None:
            clusters.append([index])
        else:
            clusters[best_cluster].append(index)
    return clusters


def deterministic_cluster_title(cluster_episodes: list[dict]) -> str:
    text = searchable_text(" ".join(item["canonical_problem"] for item in cluster_episodes))
    entities = Counter(entity for item in cluster_episodes for entity in item["entities"])
    if "github" in text and any(value in text for value in ("invite", "lời mời", "organization", "member")):
        return "Không nhận được GitHub Organization invite"
    if "vlearn" in text and any(value in text for value in ("đăng nhập", "credential", "tài khoản")):
        return "Không đăng nhập được VLearn"
    if any(value in text for value in ("xin nghỉ", "điểm danh", "vắng")):
        return "Xin nghỉ học và cập nhật điểm danh"
    if "role" in text and "discord" in text:
        return "Chưa nhận đúng Discord role"
    if "slide" in text or "bài giảng" in text:
        return "Không tìm thấy nội dung hoặc slide bài giảng"
    if entities:
        return f"Vướng mắc liên quan {entities.most_common(1)[0][0]}"
    canonical = cluster_episodes[0]["canonical_problem"].strip()
    return canonical[:96] + ("…" if len(canonical) > 96 else "")


def llm_cluster_title(cluster_episodes: list[dict], fallback: str, model: str) -> str:
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        return fallback
    examples = "\n".join(
        f"- {item['canonical_problem'][:240]}" for item in cluster_episodes[:5]
    )
    prompt = (
        "Đặt một tên pain point tiếng Việt cụ thể, tối đa 12 từ. "
        "Không thêm giải thích và không suy diễn ngoài ví dụ.\n" + examples
    )
    request = Request(
        "https://openrouter.ai/api/v1/chat/completions",
        data=json.dumps(
            {
                "model": model,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.1,
                "max_tokens": 60,
            }
        ).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "HTTP-Referer": "http://localhost:3000",
            "X-OpenRouter-Title": "Kute Dataset Pipeline",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=25) as response:
            body = json.loads(response.read().decode("utf-8"))
        title = normalize_text(body["choices"][0]["message"]["content"]).strip('"“”')
        return title[:120] if title else fallback
    except Exception:
        return fallback


def summarize_clusters(
    episodes: list[dict],
    *,
    threshold: float = 0.38,
    use_llm_labels: bool = False,
    llm_model: str = "google/gemma-4-26b-a4b-it:free",
) -> list[dict]:
    raw_clusters = cluster_episodes(episodes, threshold)
    mergeable_titles = {
        "Không nhận được GitHub Organization invite",
        "Không đăng nhập được VLearn",
        "Xin nghỉ học và cập nhật điểm danh",
        "Chưa nhận đúng Discord role",
        "Không tìm thấy nội dung hoặc slide bài giảng",
    }
    merged_clusters: list[list[int]] = []
    merge_indexes: dict[str, int] = {}
    for members in raw_clusters:
        provisional_title = deterministic_cluster_title([episodes[index] for index in members])
        if provisional_title in mergeable_titles and provisional_title in merge_indexes:
            merged_clusters[merge_indexes[provisional_title]].extend(members)
            continue
        if provisional_title in mergeable_titles:
            merge_indexes[provisional_title] = len(merged_clusters)
        merged_clusters.append(list(members))

    summaries: list[dict] = []
    for cluster_number, members in enumerate(merged_clusters, start=1):
        values = [episodes[index] for index in members]
        cluster_id = f"painpoint-{cluster_number:04d}"
        for episode in values:
            episode["painpoint_cluster_id"] = cluster_id
        fallback_title = deterministic_cluster_title(values)
        title = (
            llm_cluster_title(values, fallback_title, llm_model)
            if use_llm_labels
            else fallback_title
        )
        response_times = [
            item["time_to_first_answer"]
            for item in values
            if item["time_to_first_answer"] is not None
        ]
        unresolved_count = sum(
            item["resolution_status"] in {"unresolved", "unclear"} for item in values
        )
        known_resolutions = [item["answer_summary"] for item in values if item["answer_summary"]]
        all_dates = [parse_datetime(item["start_time"]) for item in values]
        summaries.append(
            {
                "painpoint_cluster_id": cluster_id,
                "painpoint_title": title,
                "episode_count": len(values),
                "unique_reporters": len({item["reporter_key"] for item in values}),
                "first_seen": min(all_dates).isoformat(),
                "last_seen": max(all_dates).isoformat(),
                "unresolved_rate": round(unresolved_count / len(values), 3),
                "median_time_to_first_answer": (
                    int(statistics.median(response_times)) if response_times else None
                ),
                "representative_examples": [
                    item["canonical_problem"] for item in values[:3]
                ],
                "known_resolution": known_resolutions[0][:600] if known_resolutions else None,
                "affected_area": Counter(item["product_area"] for item in values).most_common(1)[0][0],
                "evidence_rows": sorted({row for item in values for row in item["source_rows"]}),
                "embedding_method": "tfidf_word_char_hybrid",
            }
        )
    summaries.sort(key=lambda item: (item["episode_count"], item["unique_reporters"]), reverse=True)
    return summaries


def write_jsonl(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def build_summary(messages: list[dict], episodes: list[dict], painpoints: list[dict]) -> dict:
    exact_problem_counts_all = Counter(
        message["exact_normalized_hash"]
        for message in messages
        if message["is_problem"] and message["exact_normalized_hash"]
    )
    exact_problem_counts_user = Counter(
        message["exact_normalized_hash"]
        for message in messages
        if message["is_problem"]
        and not message["is_bot"]
        and message["exact_normalized_hash"]
    )
    return {
        "generated_at": datetime.now().isoformat(),
        "messages": len(messages),
        "channels": len({item["channel_key"] for item in messages}),
        "unique_reporters": len({item["reporter_key"] for item in messages}),
        "issue_episodes": len(episodes),
        "painpoint_clusters": len(painpoints),
        "flags": {
            flag: sum(bool(item[flag]) for item in messages)
            for flag in (
                "is_dot_noise", "is_greeting", "is_acknowledgement", "is_bot",
                "is_question", "is_problem", "has_attachment", "is_attachment_only",
            )
        },
        "same_author_near_duplicates": sum(
            item["same_author_duplicate_of"] is not None for item in messages
        ),
        "exact_user_problem_duplicate_rows": sum(
            count - 1 for count in exact_problem_counts_user.values() if count > 1
        ),
        "exact_problem_duplicate_rows_including_bots": sum(
            count - 1 for count in exact_problem_counts_all.values() if count > 1
        ),
        "resolved_episodes": sum(item["resolution_status"] == "resolved" for item in episodes),
        "unresolved_or_unclear_episodes": sum(
            item["resolution_status"] in {"unresolved", "unclear"} for item in episodes
        ),
        "top_painpoints": [
            {
                "painpoint_title": item["painpoint_title"],
                "episode_count": item["episode_count"],
                "unique_reporters": item["unique_reporters"],
                "unresolved_rate": item["unresolved_rate"],
            }
            for item in painpoints[:10]
        ],
    }


def run_pipeline(
    input_dir: Path,
    output_dir: Path,
    *,
    threshold: float = 0.38,
    use_llm_labels: bool = False,
    llm_model: str = "google/gemma-4-26b-a4b-it:free",
) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    messages = read_exports(input_dir)
    episodes = build_episodes(messages)
    painpoints = summarize_clusters(
        episodes,
        threshold=threshold,
        use_llm_labels=use_llm_labels,
        llm_model=llm_model,
    )
    write_jsonl(output_dir / "messages_clean.jsonl", messages)
    write_jsonl(output_dir / "issue_episodes.jsonl", episodes)
    write_jsonl(output_dir / "painpoint_summary.jsonl", painpoints)
    summary = build_summary(messages, episodes, painpoints)
    (output_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Build non-destructive Discord context datasets.")
    parser.add_argument("--input", type=Path, required=True, help="Directory containing .xlsx exports")
    parser.add_argument("--output", type=Path, required=True, help="Directory for derived JSONL files")
    parser.add_argument("--cluster-threshold", type=float, default=0.38)
    parser.add_argument("--llm-labels", action="store_true")
    parser.add_argument(
        "--llm-model",
        default=os.getenv("OPENROUTER_MODEL", "google/gemma-4-26b-a4b-it:free"),
    )
    args = parser.parse_args()
    summary = run_pipeline(
        args.input,
        args.output,
        threshold=args.cluster_threshold,
        use_llm_labels=args.llm_labels,
        llm_model=args.llm_model,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
